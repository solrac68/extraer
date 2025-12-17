from enum import Enum
import os
from pathlib import Path
import zipfile
import shutil
from openpyxl import load_workbook
from pathlib import Path
import re
import logging
from logging.handlers import RotatingFileHandler

FILEXLSX = "C:/Proyectos/Compensar/Elyon/Contingencia/Extractor/Data/Input/TESTCONTINGENCIAslimDESDECERO.xlsx"
INPUT = "C:/Proyectos/Compensar/Elyon/Contingencia/Extractor/Data/Input/Facturas_test"
OUTPUT = "C:/Proyectos/Compensar/Elyon/Contingencia/Extractor/Data/Output"
COLUMNA_RUTA_DIRECTORIO_ORIGEN = 52
COLUMNA_RUTA_FACTURA = 53
COLUMNA_RUTA_ESTADO = 54
COLUMNA_NIT_DESTINO = 1
COLUMNA_DIRECTORIO_DESTINO = 2
MAX_FILAS_HOJA = 50


logger = logging.getLogger("SuperUnzip")
logger.setLevel(logging.INFO)

manejador = RotatingFileHandler(
    'logSWO.log', 
    maxBytes=20*1024*1024, 
    backupCount=3
)

formato = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
manejador.setFormatter(formato)

logger.addHandler(manejador)

class Estado(Enum):
    DESEMPAQUETADO_COMPLETO = 1
    FACTURA_ENCONTRADA = 2
    CARPETA_CREADA_EXITO = 3
    ERROR_NO_HAY_RUTA_NIT = 4
    ERROR_DESEMPAQUETADO_PARCIAL = 5
    ERROR_DESEMPAQUETANDO_TODOS_ARCHIVOS = 6
    ERROR_ENCONTRANDO_FACTURA = 7
    ERROR_CREANDO_CARPETA = 8
    ERROR_NO_HAY_ARCHIVOS_ZIP = 9
    DIRECTORIO_DESTINO_EXISTE = 10

def crearcarpeta(carpeta) -> Estado:
    try:
        directorio = Path(carpeta)
        directorio.mkdir(parents=True, exist_ok=True)
        log = f"Directorio {OUTPUT} creado o ya existía."
        logger.info(log)
        return Estado.CARPETA_CREADA_EXITO
    except OSError as e:
        log = f"Error al crear directorios anidados: {e}"
        logger.error(log)
        return Estado.ERROR_CREANDO_CARPETA

def desempaquetar_archivos_zip(directorio_origen: str, directorio_destino: str) -> tuple[Estado, str, list[str],str]:
    """
    Busca archivos .zip recursivamente en el directorio de origen y los extrae
    en subdirectorios dentro del directorio de destino.
    
    :param directorio_origen: Ruta de la carpeta donde buscar los .zip.
    :param directorio_destino: Ruta de la carpeta donde se extraerán los contenidos.
    """
    
    path_origen = Path(directorio_origen)
    path_destino = Path(directorio_destino)
    extension_buscada = '.zip'

    log = f"Buscando archivos {extension_buscada} en: {path_origen}"
    logger.info(log)

    archivos_desempaquetados = 0
    
    # 2. Buscar archivos .zip de forma recursiva (**)
    # El uso de glob() con generador garantiza eficiencia en memoria
    error_extrayendo = False
    directorios_destino = []
    for archivo_zip in path_origen.glob(f'**/*{extension_buscada}'):
        # Solo procesamos si realmente es un archivo
        if not archivo_zip.is_file():
            continue
            
        # 3. Determinar el directorio de extracción individual
        # Usamos .stem para obtener el nombre del archivo sin la extensión (.zip)
        # Esto crea una subcarpeta de destino con el nombre del zip
        nombre_base = archivo_zip.stem
        ruta_extraccion = Path(path_destino) / nombre_base
        

        log = f"Procesando: {archivo_zip.name}"
        logger.info(log)
        
        try:
            # Creamos el subdirectorio de extracción
            ruta_extraccion.mkdir(parents=True, exist_ok=True)

            logger.info(f"Ruta extracción: {ruta_extraccion}")

            # 4. Desempaquetar el archivo
            with zipfile.ZipFile(archivo_zip, 'r') as zip_ref:
                # El método extractall hace la descompresión
                zip_ref.extractall(ruta_extraccion)
            
            logger.info(f"Éxito: Contenido extraído a -> {ruta_extraccion.name}/")
            directorios_destino.append(ruta_extraccion)
            archivos_desempaquetados += 1
            
        except zipfile.BadZipFile:
            error = f"Error: El archivo {archivo_zip.name} no es un archivo ZIP válido o está corrupto."
            logger.error(error)
            directorios_destino.append(error)
     
            shutil.rmtree(ruta_extraccion, ignore_errors=True)
            error_extrayendo = True
        except Exception as e:
            error = f"Error inesperado al descomprimir {archivo_zip.name}: {e}"
            logger.error(error)
            directorios_destino.append(error)
            error_extrayendo = True

    print(f"Proceso finalizado. Total de archivos ZIP desempaquetados: {archivos_desempaquetados}")
    
    if archivos_desempaquetados == 0 and error_extrayendo == False : return (Estado.ERROR_NO_HAY_ARCHIVOS_ZIP,directorio_origen,directorios_destino,directorio_destino)
    if archivos_desempaquetados > 0 and error_extrayendo == True : return (Estado.ERROR_DESEMPAQUETADO_PARCIAL,directorio_origen,directorios_destino,directorio_destino)
    if archivos_desempaquetados > 0 and error_extrayendo == False : return (Estado.DESEMPAQUETADO_COMPLETO,directorio_origen,directorios_destino,directorio_destino)

def desempaquetaarchivoszip(nit) -> tuple[Estado, str, list[str],str]:
    directorio_origen = Path(INPUT) / nit
    directorio_destino = Path(OUTPUT) / nit
    
    if not(os.path.exists(directorio_origen)) or not(os.path.isdir(directorio_origen)):
        logger.info(f"La carpeta {directorio_origen} no se encontró.")
        return (Estado.ERROR_NO_HAY_RUTA_NIT,str(directorio_origen),[],directorio_destino)
    
    estado = crearcarpeta(OUTPUT)
    if estado == Estado.ERROR_CREANDO_CARPETA: return (Estado.ERROR_CREANDO_CARPETA,"",[],directorio_destino)
        
    if os.path.exists(directorio_destino):
        logger.info(f"La carpeta {directorio_destino} ya existe.")
        return (Estado.DIRECTORIO_DESTINO_EXISTE,str(directorio_origen),[],directorio_destino)
    
    
    estado = crearcarpeta(directorio_destino)
    if estado == Estado.ERROR_CREANDO_CARPETA: return (Estado.ERROR_CREANDO_CARPETA,"",[],directorio_destino)

    return desempaquetar_archivos_zip(directorio_origen,directorio_destino)


def buscar_factura(pattern, directorio_destino) -> tuple[Estado, str]:
    # rglob('*') devuelve un generador de objetos Path para archivos y directorios
    for item_path in directorio_destino.rglob('*'):
        if item_path.is_dir():
            logger.info(f"Buscando el directorio {item_path.name} en el directorio destino {directorio_destino}")
            match = re.search(pattern, item_path.name)
            if match: 
                return (Estado.FACTURA_ENCONTRADA, str(item_path))
            
    for item_path in directorio_destino.rglob('*'):
        if item_path.is_file():
            logger.info(f"Buscando el archivo {item_path.name} en el directorio destino {directorio_destino}")
            match = re.search(pattern, item_path.name)
            if match: 
                return (Estado.FACTURA_ENCONTRADA, str(item_path))

    return (Estado.ERROR_ENCONTRANDO_FACTURA,"No se encontro factura")


def main():
    try:
        workbook = load_workbook(FILEXLSX, read_only=False)
    except FileNotFoundError:
        logger.error(f"El archivo {FILEXLSX} no se encontró.")
        exit()

    # 2. Seleccionar la hoja (Sheet)
    # Itera sobre los nombres de las hojas o usa el nombre directamente
    #nombres_de_hojas = workbook.sheetnames
    #print(nombres_de_hojas)
    sheet = workbook["base"]
    indice = workbook["Indice"]
    contador_directorio_destino = indice.cell(row=2, column=3).value
    nombre_hoja_directorio = f"Directorios_{contador_directorio_destino}"
    if nombre_hoja_directorio in workbook.sheetnames:
        sheet_directorios_destino = workbook[nombre_hoja_directorio]
    else:
        sheet_directorios_destino = workbook.create_sheet(f"Directorios_{contador_directorio_destino}")
        sheet_directorios_destino.cell(row=1, column=1, value="Nit")
        sheet_directorios_destino.cell(row=1, column=2, value="Ubicacion")


    # 3. Iterar sobre las filas de la hoja
    # La magia de la eficiencia está aquí: 
    # la iteración 'lee' las filas del archivo una por una
    logger.info(f"Comenzando a procesar la hoja: {sheet.title}...")
    contador_filas = indice.cell(row=2, column=1).value
    contador_filas_destino = indice.cell(row=2, column=2).value


    rows_generator = sheet.rows

    for _ in range(contador_filas):
        header = next(rows_generator)

    logger.info(f"Encabezados saltados (no procesados): {[cell.value for cell in header]}")

    for row in rows_generator:
        # 'row' es una tupla de objetos 'Cell'
        # Para obtener los valores de las celdas en la fila:
        valores_de_fila = [cell.value for cell in row]
        
        NIT = str(valores_de_fila[0]).strip()

        (estadodesempaquetado, directorio_origen, directorios_destino,directorio_destino) = desempaquetaarchivoszip(NIT)
        
        prefijo = str(valores_de_fila[2]).strip() 
        factura = str(valores_de_fila[3]).strip()
        
        pattern = f"\w*{factura}" if prefijo == "None" else f"{prefijo}\w*{factura}"
    
        logger.info(f"patron de busqueda: {pattern}")

        estadodesempaquetado, ubicacion_factura = buscar_factura(pattern, directorio_destino)
        
        sheet.cell(row=contador_filas+1, column=COLUMNA_RUTA_DIRECTORIO_ORIGEN, value=str(directorio_origen))
        sheet.cell(row=contador_filas+1, column=COLUMNA_RUTA_ESTADO, value=estadodesempaquetado.value)
        sheet.cell(row=contador_filas+1, column=COLUMNA_RUTA_FACTURA, value=ubicacion_factura)
        
        for celda_destino in directorios_destino:
            sheet_directorios_destino.cell(row=contador_filas_destino + 1, column=COLUMNA_NIT_DESTINO, value=str(NIT))
            sheet_directorios_destino.cell(row=contador_filas_destino + 1, column=COLUMNA_DIRECTORIO_DESTINO, value=str(celda_destino))
            contador_filas_destino += 1
            if contador_filas_destino % MAX_FILAS_HOJA == 0:
                contador_directorio_destino += 1
                contador_filas_destino = 1
                sheet_directorios_destino = workbook.create_sheet(f"Directorios_{contador_directorio_destino}")
                sheet_directorios_destino.cell(row=1, column=1, value="Nit")
                sheet_directorios_destino.cell(row=1, column=2, value="Ubicacion")

        contador_filas += 1
        
        indice.cell(row=2, column=1, value=contador_filas)
        indice.cell(row=2, column=2, value=contador_filas_destino)
        indice.cell(row=2, column=3, value=contador_directorio_destino)

        try:
            workbook.save(FILEXLSX)
            logger.info(f"Éxito: El archivo '{FILEXLSX}' ha sido guardado con los nuevos valores.")
        except Exception as e:
            logger.error(f"Error al guardar el archivo: {e}")
        
        log = f"Procesamiento completado. Se leyeron {contador_filas} filas."
        print(log)
        logger.info(log)
    
    print("-------PROCESO FINALIZADO---------------")
    logger.info("-------PROCESO FINALIZADO---------------")

if __name__ == "__main__":
    main()